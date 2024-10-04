import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from tkinter import filedialog  # Para abrir el diálogo de archivos
import xlsxwriter
import openpyxl  # Para leer archivos Excel

class HojaCalculo:
    def __init__(self, root):
        self.root = root
        self.root.title("Hoja de Cálculo Dinámica")

        # Aplicar tema a ttk
        self.style = ttk.Style()
        self.style.theme_use('clam')

        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)

        # Configurar frame principal
        self.frame = ttk.Frame(self.root, padding="0")
        self.frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Crear botones y layout
        self.agregar_fila_btn = ttk.Button(self.frame, text="Agregar Fila", command=self.agregar_fila)
        self.agregar_fila_btn.grid(row=0, column=0, padx=50, pady=10, sticky=tk.W)

        self.reducir_fila_btn = ttk.Button(self.frame, text="Reducir Fila", command=self.reducir_fila)
        self.reducir_fila_btn.grid(row=0, column=1, padx=50, pady=10, sticky=tk.W)

        self.agregar_columna_btn = ttk.Button(self.frame, text="Agregar Columna", command=self.agregar_columna)
        self.agregar_columna_btn.grid(row=0, column=2, padx=50, pady=10, sticky=tk.W)

        self.reducir_columna_btn = ttk.Button(self.frame, text="Reducir Columna", command=self.reducir_columna)
        self.reducir_columna_btn.grid(row=0, column=3, padx=50, pady=10, sticky=tk.W)

        self.guardar_excel_btn = ttk.Button(self.frame, text="Guardar Excel", command=self.guardar_excel)
        self.guardar_excel_btn.grid(row=0, column=4, padx=50, pady=10, sticky=tk.W)

        # Botón para cargar archivo Excel
        self.cargar_excel_btn = ttk.Button(self.frame, text="Cargar Excel", command=self.cargar_excel)
        self.cargar_excel_btn.grid(row=0, column=5, padx=50, pady=10, sticky=tk.W)

        # Frame para operaciones
        self.operaciones_frame = ttk.LabelFrame(self.frame, text="Operaciones Aritméticas", padding="10 10 10 10")
        self.operaciones_frame.grid(row=1, column=0, columnspan=6, padx=50, pady=5, sticky=(tk.W, tk.E))

        ttk.Label(self.operaciones_frame, text="Celdas:").grid(row=0, column=0, padx=50, pady=5)

        self.fila1_entry = ttk.Entry(self.operaciones_frame, width=3)
        self.fila1_entry.grid(row=0, column=1)
        ttk.Label(self.operaciones_frame, text=",").grid(row=0, column=2)
        self.col1_entry = ttk.Entry(self.operaciones_frame, width=3)
        self.col1_entry.grid(row=0, column=3)

        self.operacion_var = tk.StringVar()
        self.operacion_var.set("+")  # valor predeterminado
        operaciones_menu = ttk.OptionMenu(self.operaciones_frame, self.operacion_var, "+", "+", "-", "*", "/")
        operaciones_menu.grid(row=0, column=4, padx=50, pady=5)

        self.fila2_entry = ttk.Entry(self.operaciones_frame, width=3)
        self.fila2_entry.grid(row=0, column=5)
        ttk.Label(self.operaciones_frame, text=",").grid(row=0, column=6)
        self.col2_entry = ttk.Entry(self.operaciones_frame, width=3)
        self.col2_entry.grid(row=0, column=7)

        self.resultado_lbl = ttk.Label(self.operaciones_frame, text="Resultado: ")
        self.resultado_lbl.grid(row=0, column=8, padx=50, pady=5)

        self.calcular_btn = ttk.Button(self.operaciones_frame, text="Calcular", command=self.realizar_operacion)
        self.calcular_btn.grid(row=0, column=9, padx=50, pady=5)

        # Frame para las celdas de la hoja de cálculo con scrollbars
        self.celdas_frame = ttk.Frame(self.frame, padding="0")
        self.celdas_frame.grid(row=2, column=0, columnspan=6, sticky=(tk.W, tk.E))

        self.canvas = tk.Canvas(self.celdas_frame)
        self.scroll_y = ttk.Scrollbar(self.celdas_frame, orient="vertical", command=self.canvas.yview)
        self.scroll_x = ttk.Scrollbar(self.celdas_frame, orient="horizontal", command=self.canvas.xview)

        self.scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

        self.scroll_frame = ttk.Frame(self.canvas)

        self.scroll_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))

        self.canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.scroll_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))

        self.celdas = []

        # Encabezados de columnas
        self.encabezados = []  
        for col, letra in enumerate(self.encabezados):
            encabezado = ttk.Label(self.scroll_frame, text=letra, font=("Arial", 12, "bold"))
            encabezado.grid(row=0, column=col, padx=5, pady=5)

    def agregar_fila(self):
        fila = len(self.celdas)
        num_columnas = len(self.celdas[0]) if self.celdas else len(self.encabezados)  
        nueva_fila = [ttk.Entry(self.scroll_frame, width=10) for _ in range(num_columnas)]
        for entrada in nueva_fila:
            entrada.insert(0, "")  # Dejar las celdas vacías
            entrada.grid(row=fila + 1, column=nueva_fila.index(entrada), padx=5, pady=5) 

        self.celdas.append(nueva_fila)

    def reducir_fila(self):
        if self.celdas:
            fila_a_borrar = len(self.celdas) - 1  
            for entrada in self.celdas[fila_a_borrar]:
                entrada.destroy() 
            self.celdas.pop()  

    def agregar_columna(self):
        if not self.celdas:
            self.agregar_fila()
        else:
            for fila in range(len(self.celdas)):
                nueva_celda = ttk.Entry(self.scroll_frame, width=10)
                nueva_celda.insert(0, "")  # Dejar la celda vacía
                nueva_celda.grid(row=fila + 1, column=len(self.celdas[fila]), padx=5, pady=5)
                self.celdas[fila].append(nueva_celda)

            nueva_letra = chr(ord(self.encabezados[-1]) + 1) 
            self.encabezados.append(nueva_letra)
            encabezado = ttk.Label(self.scroll_frame, text=nueva_letra, font=("Arial", 12, "bold"))
            encabezado.grid(row=0, column=len(self.encabezados) - 1, padx=5, pady=5)

    def reducir_columna(self):
        if self.celdas:
            for fila in self.celdas:
                entrada = fila.pop() 
                entrada.destroy()  

            self.encabezados.pop() 

    def realizar_operacion(self):
        try:
            fila1 = int(self.fila1_entry.get()) - 1  
            col1 = ord(self.col1_entry.get().upper()) - ord('A')  
            fila2 = int(self.fila2_entry.get()) - 1  
            col2 = ord(self.col2_entry.get().upper()) - ord('A') 

            # Validar índices
            if fila1 < 0 or fila2 < 0 or col1 < 0 or col2 < 0:
                raise IndexError("Los índices deben ser mayores que cero.")
            if fila1 >= len(self.celdas) or fila2 >= len(self.celdas):
                raise IndexError("Los índices de fila están fuera del rango de la hoja de cálculo.")
            if col1 >= len(self.celdas[0]) or col2 >= len(self.celdas[0]):
                raise IndexError("Los índices de columna están fuera del rango de la hoja de cálculo.")

            valor1 = float(self.celdas[fila1][col1].get() or 0)  
            valor2 = float(self.celdas[fila2][col2].get() or 0)  

            if self.operacion_var.get() == "+":
                resultado = valor1 + valor2
            elif self.operacion_var.get() == "-":
                resultado = valor1 - valor2
            elif self.operacion_var.get() == "*":
                resultado = valor1 * valor2
            elif self.operacion_var.get() == "/":
                resultado = valor1 / valor2 if valor2 != 0 else "Error: División por cero"

            self.resultado_lbl.config(text=f"Resultado: {resultado}")

        except ValueError:
            messagebox.showerror("Error", "Por favor, ingrese valores válidos.")
        except IndexError as e:
            messagebox.showerror("Error", str(e))

    def cargar_excel(self):
        archivo = filedialog.askopenfilename(filetypes=[("Archivos de Excel", "*.xlsx")])
        if archivo:
            self.limpiar_celdas()
            wb = openpyxl.load_workbook(archivo)
            hoja = wb.active

            for i, fila in enumerate(hoja.iter_rows(values_only=True)):
                nueva_fila = []
                for j, valor in enumerate(fila):
                    if i == 0:
                        if j >= len(self.encabezados):
                            self.encabezados.append(chr(ord('A') + j))
                            encabezado = ttk.Label(self.scroll_frame, text=self.encabezados[j], font=("Arial", 12, "bold"))
                            encabezado.grid(row=0, column=j, padx=5, pady=5)
                    # Crear las entradas de datos para cada fila
                    entrada = ttk.Entry(self.scroll_frame, width=10)
                    entrada.insert(0, valor if valor is not None else "")  
                    entrada.grid(row=i + 1, column=j, padx=5, pady=5)
                    nueva_fila.append(entrada)
                # Añadir la fila a la lista de celdas
                self.celdas.append(nueva_fila)


    def limpiar_celdas(self):
        for fila in self.celdas:
            for entrada in fila:
                entrada.destroy()
        self.celdas.clear()
        self.encabezados = []  

    def guardar_excel(self):
        archivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Excel", "*.xlsx")])
        if archivo:
            wb = xlsxwriter.Workbook(archivo)
            hoja = wb.add_worksheet()

            # Escribir encabezados
            for col, letra in enumerate(self.encabezados):
                hoja.write(0, col, letra)

            # Escribir datos
            for i, fila in enumerate(self.celdas):
                for j, entrada in enumerate(fila):
                    valor = entrada.get()
                    hoja.write(i + 1, j, valor)

            wb.close()
            messagebox.showinfo("Éxito", "Archivo guardado exitosamente.")

if __name__ == "__main__":
    root = tk.Tk()
    app = HojaCalculo(root)
    root.mainloop()
